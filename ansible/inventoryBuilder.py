# -*- coding: utf-8 -*-
"""
Created on Sat Aug 20 10:38:50 2022

@author: alexi
"""

from openpyxl import load_workbook
import re
from optparse import OptionParser
import ipaddress

class Computer:
    def __init__(self, name, osType):
        self.name = name
        self.osType = osType
        

parser = OptionParser()
parser.add_option("-i", "--in", action="store", type="string",dest="inputName",
                  default="Network.xlsx", help="Network File Name")
parser.add_option("-o", "--out", action="store", type="string",dest="outputName",
                  default="inventory.yaml", help="Inventory Output File Name")
(options, args) = parser.parse_args()


def printHostGroups(exerciseDomains,teamHosts,teamName):
    for domain in exerciseDomains:
        if "hybrid" not in exerciseDomains[domain]['os']:
            print(f"    {domain}_{teamName}_{exerciseDomains[domain]['os']}:\n"
                  f"      vars:\n"
                  f"        passwd_pfx: \"{teamHosts[domain]['prefix']}\"")
            if exerciseDomains[domain]['os'] == "windows":
                print(f"        domain_name: {domain[:3]}{teamName[:3]}")
                
                windowsGroup.append(f"{domain}_{teamName}_windows")
            print("      hosts:")
            for host in teamHosts[domain]['hosts']:
                print(f"        {host.name}.{domain}.{teamName}:")
        else:
            print(f"    {domain}_{teamName}_windows:")
            windowsGroup.append(f"{domain}_{teamName}_windows:")
            print(f"      vars:\n"
                  f"        passwd_pfx: \"{teamHosts[domain]['prefix']}\"\n"
                  f"        domain_name: {domain[:3]}{teamName[:3]}\n"
                  f"      hosts:")
            for host in teamHosts[domain]['hosts']:
                if host.osType == "windows":
                    print(f"        {host.name}.{domain}.{teamName}:")
                    
            print(f"    {domain}_{teamName}_linux:\n"
                  f"      vars:\n"
                  f"        passwd_pfx: \"{teamHosts[domain]['prefix']}\"\n"
                  f"      hosts:")
            for host in teamHosts[domain]['hosts']:
                if host.osType == "linux":
                    print(f"        {host.name}.{domain}.{teamName}:")    

def rebuildExerciseDomains(exerciseDomains,ipCol,domainCol):
    for domains in domainCol:
        if domains.value is not None and re.match("host.[a-z]+.civ",domains.value):
            domainName = domains.value.split('.')[1]
            network = ipaddress.ip_network(f"{ipCol[domains.row].value}/24",
                                           strict=False)
            exerciseDomains[domainName]['networkAddress'] = network
    return exerciseDomains

def rebuildTeamHosts(exerciseDomains, teamHosts, domainCol, osCol, hostCol, 
                     infoCol):
    for domain in teamHosts:
        teamHosts[domain]['hosts'] = []
        
    for host in hostCol:
        if host.value is not None and host.value.islower():
            hostIP = ipaddress.ip_address(infoCol[host.row-1].value)
            for domainName in exerciseDomains:
                if hostIP in exerciseDomains[domainName]['networkAddress']:
                    os = osCol[host.row-1].value
                    #This determins the os type of host 
                    if "Win" in os:
                        teamHosts[domainName]['hosts'].append(
                            Computer(host.value,"windows"))
                    else:
                        teamHosts[domainName]['hosts'].append(
                            Computer(host.value,"linux"))
    buildPrefix(domainCol, infoCol, teamHosts)

def buildPrefix(domainCol, infoCol, teamHosts):
    for prefix in infoCol:
        if prefix.value is not None and len(prefix.value)==3:
            domainOfPrefix = domainCol[prefix.row-1].value.split('.')[1]
            #Checks if the prefix has a special character that yaml can't understand
            if '"' in prefix.value:
                idx = prefix.value.index('"')
                teamHosts[domainOfPrefix]['prefix'] = (prefix.value[:idx]+
                                                    "\\"+prefix.value[idx:])
            else:
                teamHosts[domainOfPrefix]['prefix'] = prefix.value
                
workbook = load_workbook(filename=options.inputName)

sheet = workbook.active

B = sheet["B"]


#Build two dictionaries
#exerciseDomains will hold all of the domains inside the exercise, and the cell color of said domain
#team hosts will hold each domain for each team, it will get created alongside exerciseDomains
exerciseDomains = {}
teamHosts = {}
numTeam = int((sheet.max_column-3)/2)
for domains in B:
    if domains.value is not None and re.match("host.[a-z]+.civ",domains.value):
        domainName = domains.value.split('.')[1]
        if domainName not in exerciseDomains:
            exerciseDomains[domainName] = {}
            teamHosts[domainName] = {}
            teamHosts[domainName]['hosts'] = [] 
            exerciseDomains[domainName]['fqdn'] = domains.value
            #exerciseDomains[domainName]['networkAddress']
            
            row = domains.row
            ipColIterator = "E"
            
            #for x in range(numTeam):
            ipCol = sheet[ipColIterator]

            network = ipaddress.ip_network(f"{ipCol[row].value}/24", 
                                            strict=False)
            #print(network)
            exerciseDomains[domainName]['networkAddress'] = network
                # i = ord(ipColIterator[0])
                # i += 2
                # ipColIterator = chr(i)

osType = sheet["C"]
hostColIterator = "D"
infoColIterator = "E"
#for team in range(numTeam):
    
hostCol = sheet[hostColIterator]
teamName = hostCol[0].value
infoCol = sheet[infoColIterator]
windowsGroup = []


for host in hostCol:
    if host.value is not None and host.value.islower():
        hostIP = ipaddress.ip_address(infoCol[host.row-1].value)
        for domainName in exerciseDomains:
            if hostIP in exerciseDomains[domainName]['networkAddress']:
                os = osType[host.row-1].value
                #This determins the os type of host 
                if "Win" in os:

                    teamHosts[domainName]['hosts'].append(
                        Computer(host.value,"windows"))
                else:
                    teamHosts[domainName]['hosts'].append(
                        Computer(host.value,"linux"))
                    
                #This marks the type of domain
                if 'os' in exerciseDomains[domainName] and hostColIterator == "D":
                    if exerciseDomains[domainName] != 'hybrid':
                        if ("Win" in os and exerciseDomains[domainName]['os'] != "windows") or ("Win" not in os and exerciseDomains[domainName]['os'] == "windows"):
                            exerciseDomains[domainName]['os'] = "hybrid"
                elif 'os' not in exerciseDomains[domainName] and hostColIterator == "D":
                    if "Win" in os:
                        exerciseDomains[domainName]['os'] = "windows"
                    else:
                        exerciseDomains[domainName]['os'] = "linux"

buildPrefix(B, infoCol, teamHosts)


print("all:\n  children:")
for x in range(int((sheet.max_column-3)/2)):
#for x in range(2):
    printHostGroups(exerciseDomains,teamHosts,teamName.lower())
    
    if x != int(((sheet.max_column-3)/2)-1):
        i = ord(hostColIterator[0])
        i += 2
        hostColIterator = chr(i)
        infoColIterator = chr(i + 1)
        
        hostCol = sheet[hostColIterator]
        teamName = hostCol[0].value
        infoCol = sheet[infoColIterator]
        
        rebuildExerciseDomains(exerciseDomains, infoCol, B)
        rebuildTeamHosts(exerciseDomains, teamHosts, B, osType, hostCol, 
                         infoCol)     